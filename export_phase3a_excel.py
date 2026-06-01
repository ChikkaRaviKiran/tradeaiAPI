"""Export Phase-3a actual (±₹6k, fine-grid) trade results to Excel.
Source: phase3a_actual_6k.py run on AWS Lightsail (DhanHQ-sourced option_candles).
"""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"c:\TradeAI\research\reports\phase3a_actual_6k_trades.xlsx"

# ── 29 trades (latest run with full Dhan refetch, Apr 14 → May 29, 2026) ──
TRADES = [
    # (n, date, wd, sym, strategy, entry, exit, sl_pct, why, pnl_rs)
    (1,  date(2026,4,15), "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TP",    6069),
    (2,  date(2026,4,16), "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TP",    6420),
    (3,  date(2026,4,17), "Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, "TIME",  2223),
    (4,  date(2026,4,20), "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TIME",   302),
    (5,  date(2026,4,21), "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TIME",  4144),
    (6,  date(2026,4,22), "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TP",    6057),
    (7,  date(2026,4,23), "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TP",    6123),
    (8,  date(2026,4,24), "Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, "TIME",  1940),
    (9,  date(2026,4,27), "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TP",    6064),
    (10, date(2026,4,28), "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TP",    6123),
    (11, date(2026,4,29), "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TIME",  3522),
    (12, date(2026,4,30), "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TP",    6198),
    (13, date(2026,5,4),  "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TP",    6357),
    (14, date(2026,5,5),  "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TP",    6298),
    (15, date(2026,5,6),  "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TIME",  2010),
    (16, date(2026,5,7),  "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TP",    6093),
    (17, date(2026,5,8),  "Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, "TIME",  5655),
    (18, date(2026,5,11), "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TIME",  5558),
    (19, date(2026,5,12), "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TIME",  3461),
    (20, date(2026,5,13), "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TP",    8514),
    (21, date(2026,5,14), "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TIME",  3930),
    (22, date(2026,5,15), "Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, "TP",    6006),
    (23, date(2026,5,18), "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TIME",  3247),
    (24, date(2026,5,19), "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TP",    6162),
    (25, date(2026,5,20), "Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, "TIME",  2487),
    (26, date(2026,5,21), "Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, "TIME",  2721),
    (27, date(2026,5,22), "Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, "TIME",  5733),
    (28, date(2026,5,25), "Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, "TIME",  4524),
    (29, date(2026,5,26), "Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, "TIME",  5899),
]

SCHEDULE = [
    # (Day, Index, Strategy, Entry, Exit, SL%, Train, Test, Full, Worst, Sharpe, W/N)
    ("Mon", "NIFTY",  "strangle+1", "09:25", "12:00", 30, 12724, 13328, 26052,  302,  2.08, "6/6"),
    ("Tue", "NIFTY",  "strangle+1", "09:25", "12:00", 50, 16565, 15522, 32087,  3461, 4.79, "6/6"),
    ("Wed", "SENSEX", "strangle+2", "10:00", "15:15", 30, 15648, 13011, 28659,  2010, 2.08, "6/6"),
    ("Thu", "SENSEX", "strangle+2", "09:45", "13:30", 30, 18741, 12744, 31485,  2721, 3.73, "6/6"),
    ("Fri", "NIFTY",  "straddle",   "09:45", "14:30", 30, 4163,  17394, 21557,  1940, 2.36, "5/5"),
]

# ── Styles ──
HDR  = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL= PatternFill("solid", fgColor="FFEB9C")
GREEN_FONT = Font(color="006100", bold=True)
RED_FONT   = Font(color="9C0006", bold=True)

wb = Workbook()

# ════════════ SHEET 1: Schedule (final recommendation) ════════════
ws1 = wb.active
ws1.title = "Schedule"
ws1["A1"] = "Phase-3a ACTUAL — Final Recommendation (±₹6,000 gates, 3 lots, one index/day)"
ws1["A1"].font = Font(bold=True, size=13, color="1F4E78")
ws1.merge_cells("A1:L1")
ws1["A2"] = "Data: DhanHQ /v2/charts/rollingoption · Window: 2026-04-14 → 2026-05-29 · Walk-forward: Train Apr15–May5, Test May6–May29"
ws1["A2"].font = Font(italic=True, color="595959", size=10)
ws1.merge_cells("A2:L2")

hdr1 = ["Day","Index","Strategy","Entry","Exit","SL %","Train ₹","Test ₹","Full ₹","Worst ₹","Sharpe","W/N"]
for c, h in enumerate(hdr1, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

for r, row in enumerate(SCHEDULE, start=5):
    for c, v in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = CENTER if c <= 6 or c == 11 or c == 12 else RIGHT
        if c in (7,8,9,10):
            cell.number_format = '"₹"#,##0;[Red]"-₹"#,##0'

# totals row
tot_train = sum(s[6] for s in SCHEDULE)
tot_test  = sum(s[7] for s in SCHEDULE)
tot_full  = sum(s[8] for s in SCHEDULE)
row_tot = 5 + len(SCHEDULE)
ws1.cell(row=row_tot, column=1, value="TOTAL").font = Font(bold=True)
ws1.cell(row=row_tot, column=7, value=tot_train).number_format = '"₹"#,##0;[Red]"-₹"#,##0'
ws1.cell(row=row_tot, column=8, value=tot_test).number_format  = '"₹"#,##0;[Red]"-₹"#,##0'
ws1.cell(row=row_tot, column=9, value=tot_full).number_format  = '"₹"#,##0;[Red]"-₹"#,##0'
for c in range(1, 13):
    ws1.cell(row=row_tot, column=c).font = Font(bold=True)
    ws1.cell(row=row_tot, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
    ws1.cell(row=row_tot, column=c).border = BORDER

widths1 = [6,9,13,8,8,7,12,12,12,12,9,7]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ════════════ SHEET 2: All trades ════════════
ws2 = wb.create_sheet("Trades")
ws2["A1"] = "Phase-3a ACTUAL — Trade-by-trade Log (25 trades)"
ws2["A1"].font = Font(bold=True, size=13, color="1F4E78")
ws2.merge_cells("A1:K1")

hdr2 = ["#","Date","WD","Index","Strategy","Entry","Exit","SL %","Exit Reason","P&L ₹","Cumulative ₹"]
for c, h in enumerate(hdr2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

cum = 0
for n, d, wd, sym, strat, en, ex, sl, why, pnl in TRADES:
    cum += pnl
    r = 3 + n
    vals = [n, d, wd, sym, strat, en, ex, sl, why, pnl, cum]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.border = BORDER
        if c == 2: cell.number_format = "yyyy-mm-dd"; cell.alignment = CENTER
        elif c in (10,11):
            cell.number_format = '"₹"#,##0;[Red]"-₹"#,##0'; cell.alignment = RIGHT
        else: cell.alignment = CENTER
    # color by exit reason
    why_cell = ws2.cell(row=r, column=9)
    pnl_cell = ws2.cell(row=r, column=10)
    if why == "TP":
        why_cell.fill = GREEN_FILL; why_cell.font = GREEN_FONT
        pnl_cell.fill = GREEN_FILL
    elif why == "SL_RS":
        why_cell.fill = RED_FILL; why_cell.font = RED_FONT
        pnl_cell.fill = RED_FILL
    elif why == "SL_PCT":
        why_cell.fill = YELLOW_FILL
        pnl_cell.fill = YELLOW_FILL
    if pnl < 0: pnl_cell.font = RED_FONT
    else:       pnl_cell.font = GREEN_FONT

# totals
total_pnl = sum(t[9] for t in TRADES)
n_wins = sum(1 for t in TRADES if t[9] > 0)
tp_n   = sum(1 for t in TRADES if t[8] == "TP")
tot_row = 3 + len(TRADES) + 1
ws2.cell(row=tot_row, column=4, value="TOTAL").font = Font(bold=True)
ws2.cell(row=tot_row, column=10, value=total_pnl).number_format = '"₹"#,##0;[Red]"-₹"#,##0'
for c in range(1, 12):
    ws2.cell(row=tot_row, column=c).font = Font(bold=True)
    ws2.cell(row=tot_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
    ws2.cell(row=tot_row, column=c).border = BORDER

widths2 = [5,12,5,9,13,8,8,7,13,12,14]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════ SHEET 3: Summary ════════════
ws3 = wb.create_sheet("Summary")
ws3["A1"] = "Phase-3a ACTUAL — Performance Summary"
ws3["A1"].font = Font(bold=True, size=13, color="1F4E78")
ws3.merge_cells("A1:B1")

apr_trades = [t for t in TRADES if t[1] <= date(2026,4,30)]
may_trades = [t for t in TRADES if t[1]  > date(2026,4,30)]
apr_pnl = sum(t[9] for t in apr_trades)
may_pnl = sum(t[9] for t in may_trades)
worst = min(t[9] for t in TRADES)
best  = max(t[9] for t in TRADES)

summary = [
    ("Data Source",          "DhanHQ /v2/charts/rollingoption (1-min OHLC)"),
    ("Backtest Window",      "2026-04-14 → 2026-05-26 (30 trading days)"),
    ("Strategies Searched",  "8 (straddle, strangle+1, strangle+2, ironfly W2/W3/W4, condor S1W3/S2W4)"),
    ("Grid Combinations",    "23,040 (8 entries × 9 exits × 4 SL% × 8 strats × 10 buckets)"),
    ("Robust Survivors",     "5,809 (profitable in BOTH train and test halves)"),
    ("Exit Gates",           "TP=+₹6,000 · SL=-₹6,000 · SL%-of-credit safety · Time"),
    ("Lots per Trade",       "3 (NIFTY 65×3=195 qty, SENSEX 20×3=60 qty)"),
    ("",""),
    ("Total Trades",         len(TRADES)),
    ("Winning Trades",       n_wins),
    ("Losing Trades",        len(TRADES) - n_wins),
    ("Win Rate",             f"{n_wins/len(TRADES)*100:.1f}%"),
    ("",""),
    ("Total P&L",            total_pnl),
    ("April P&L",            apr_pnl),
    ("May P&L",              may_pnl),
    ("Best Trade",           best),
    ("Worst Trade",          worst),
    ("Avg per Trade",        round(total_pnl/len(TRADES))),
    ("Monthly Run-rate",     round(total_pnl / 30 * 22)),  # 22 trading days/mo
    ("",""),
    ("TP Exits (+₹6k)",      tp_n),
    ("TIME Exits",           sum(1 for t in TRADES if t[8]=="TIME")),
    ("SL_RS Exits (-₹6k)",   sum(1 for t in TRADES if t[8]=="SL_RS")),
    ("",""),
    ("Peak Day Margin (NRML)", "~₹3.0 L (Thursday SENSEX)"),
    ("Peak Day Margin (MIS)",  "~₹1.8 L"),
    ("Recommended Capital",    "₹4.0 L NRML or ₹2.5 L MIS"),
]
for r, (k, v) in enumerate(summary, start=3):
    ws3.cell(row=r, column=1, value=k).font = Font(bold=True)
    cell = ws3.cell(row=r, column=2, value=v)
    if isinstance(v, (int, float)) and k not in ("Total Trades","Winning Trades","Losing Trades",
                                                  "TP Exits (+₹6k)","TIME Exits","SL_RS Exits (-₹6k)"):
        cell.number_format = '"₹"#,##0;[Red]"-₹"#,##0'
        if isinstance(v, int) and v > 0: cell.font = GREEN_FONT
        elif isinstance(v, int) and v < 0: cell.font = RED_FONT

ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 65

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Sheet 1 (Schedule):   final 5-day recommendation with stats")
print(f"  Sheet 2 (Trades):     {len(TRADES)} individual trades with cumulative P&L (color-coded)")
print(f"  Sheet 3 (Summary):    aggregate performance metrics")

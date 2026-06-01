#!/usr/bin/env python3
"""Export Phase-3a + Phase-3b trade logs to Excel — APRIL + MAY 2026 ONLY.

Skips the expensive top-5 re-sweep; just simulates the chosen schedule on
apr/may trading days. Fast (~3-5 min).
"""
import json
from collections import defaultdict
from datetime import date
from phase3a_breakdown_v2 import (parse_expiry, to_min, load_chain,
                                   LOT_SIZE, LOTS)
from strategy_definedrisk_sweep import (find_atm, get_price, build_strategies,
                                         find_days)
from export_strategy_results import (simulate_detail, hhmm, build_trade_rows,
                                      monthly_summary)

OUT = "/tmp/tradeai_aprmay_trades.xlsx"
MONTHS = (4, 5)  # April, May 2026

def load_cache_aprmay():
    near = find_days()
    cache = {}
    for (d, sym), exps in near.items():
        if d.year != 2026 or d.month not in MONTHS:
            continue
        ex, exp_s = exps[0]
        bars = load_chain(d, sym, exp_s)
        if bars:
            cache[(d, sym)] = (bars, (ex - d).days, exp_s)
    return cache

def main():
    print("Loading schedule from /tmp/strategy_definedrisk.json ...")
    with open("/tmp/strategy_definedrisk.json") as f:
        sched = json.load(f)

    print("Loading Apr+May option chains...")
    cache = load_cache_aprmay()
    print(f"  cache: {len(cache)} (date,sym) combos")

    chosen_3a = {}
    for wd, spec in sched["phase3a"].items():
        chosen_3a[wd] = {"sym": spec["sym"], "strat": spec["strat"],
                          "entry": spec["entry"], "exit": spec["exit"],
                          "sl": int(spec["sl"]),
                          "tp": (int(spec["tp"]) if spec.get("tp") not in (None,"","None") else None)}
    rows_3a = build_trade_rows(chosen_3a, cache)
    monthly_3a = monthly_summary(rows_3a)

    chosen_3b = {}
    for key, spec in sched["phase3b"].items():
        wd, sym = key.split("_")
        chosen_3b[(wd, sym)] = {"sym": sym, "strat": spec["strat"],
                                 "entry": spec["entry"], "exit": spec["exit"],
                                 "sl": int(spec["sl"]),
                                 "tp": (int(spec["tp"]) if spec.get("tp") not in (None,"","None") else None)}
    rows_3b = build_trade_rows(chosen_3b, cache)
    monthly_3b = monthly_summary(rows_3b)

    sched_def = []
    for wd in ["Mon","Tue","Wed","Thu","Fri"]:
        s = chosen_3a.get(wd)
        if s:
            sched_def.append({"phase":"3a", "weekday":wd, "index":s["sym"],
                              "strategy":s["strat"], "entry_time":s["entry"],
                              "exit_time":s["exit"], "SL_pct":s["sl"],
                              "TP_pct":s["tp"] or "none",
                              "lots":LOTS, "lot_size":LOT_SIZE[s["sym"]]})
    for (wd, sym), s in chosen_3b.items():
        sched_def.append({"phase":"3b", "weekday":wd, "index":sym,
                          "strategy":s["strat"], "entry_time":s["entry"],
                          "exit_time":s["exit"], "SL_pct":s["sl"],
                          "TP_pct":s["tp"] or "none",
                          "lots":LOTS, "lot_size":LOT_SIZE[sym]})

    print(f"\nWriting {OUT} ...")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "Schedule_Definition"

    def write_sheet(ws, rows):
        if not rows:
            ws.append(["no data"]); return
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h,"") for h in headers])
        for i, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(r.get(h,""))) for r in rows])
            col = chr(64+i) if i <= 26 else 'A'+chr(64+i-26)
            ws.column_dimensions[col].width = min(max_len+2, 40)

    write_sheet(ws, sched_def)
    write_sheet(wb.create_sheet("Phase3a_Trades"), rows_3a)
    write_sheet(wb.create_sheet("Phase3a_Monthly"), monthly_3a)
    write_sheet(wb.create_sheet("Phase3b_Trades"), rows_3b)
    write_sheet(wb.create_sheet("Phase3b_Monthly"), monthly_3b)
    wb.save(OUT)

    print("\nDone:")
    print(f"  Schedule_Definition  ({len(sched_def)} rows)")
    print(f"  Phase3a_Trades       ({len(rows_3a)} rows)  cum=₹{rows_3a[-1]['cumulative_pnl'] if rows_3a else 0:,.0f}")
    print(f"  Phase3a_Monthly      ({len(monthly_3a)} rows)")
    print(f"  Phase3b_Trades       ({len(rows_3b)} rows)  cum=₹{rows_3b[-1]['cumulative_pnl'] if rows_3b else 0:,.0f}")
    print(f"  Phase3b_Monthly      ({len(monthly_3b)} rows)")
    print(f"\nFile: {OUT}")

if __name__ == "__main__":
    main()

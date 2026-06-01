#!/usr/bin/env python3
"""Export full Phase-3a + Phase-3b trade logs to Excel.

Re-uses chosen schedule from /tmp/strategy_definedrisk.json (written by
strategy_definedrisk_sweep.py). For each scheduled trade, re-runs the
minute-by-minute simulator and records:
  - date, weekday, index, strategy, ATM strike, all leg prices at entry,
  - entry credit (points & ₹), SL/TP triggers, exit minute, exit reason,
  - exit prices, exit cost (points & ₹), net PnL (₹), running cumulative.

Writes /tmp/tradeai_strategy_full_results.xlsx with sheets:
  - Phase3a_Trades
  - Phase3b_Trades
  - Monthly_Summary_3a
  - Monthly_Summary_3b
  - Weekday_Top5
  - Schedule_Definition
"""
import json
from collections import defaultdict
from datetime import datetime, date
from phase3a_breakdown_v2 import (psql, parse_expiry, to_min, fmt_rs,
                                   load_chain, LOT_SIZE, LOTS, STRIKE_STEP)
from strategy_definedrisk_sweep import (find_atm, get_price, build_strategies,
                                         simulate, START, END, REGIME_SPLIT,
                                         find_days)

OUT = "/tmp/tradeai_strategy_full_results.xlsx"

# ----- simulate one day with FULL bookkeeping -----
def simulate_detail(bars, legs, e_min, x_min, sl_pct, tp_pct, lotsize):
    entries = []
    for k, side, qty in legs:
        e = get_price(bars, k, side, e_min)
        if e is None: return None
        entries.append((k, side, qty, e))
    credit = sum((e if q == -1 else -e) for _,_,q,e in entries)
    if credit <= 0: return None
    sl_amt = credit * sl_pct / 100.0
    tp_amt = credit * tp_pct / 100.0 if tp_pct else None
    exit_m = x_min; exit_reason = "TIME"
    exit_mtm_pts = None
    for m in range(e_min + 1, x_min + 1):
        ok = True; mtm = 0.0; cur_list = []
        for k, side, qty, e in entries:
            p = get_price(bars, k, side, m)
            if p is None: ok = False; break
            cur_list.append(p)
            if qty == -1: mtm += (e - p)
            else:         mtm += (p - e)
        if not ok: continue
        if tp_amt is not None and mtm >= tp_amt:
            exit_m, exit_reason, exit_mtm_pts = m, "TP", mtm
            exits_at_trigger = cur_list
            break
        if mtm <= -sl_amt:
            exit_m, exit_reason, exit_mtm_pts = m, "SL", mtm
            exits_at_trigger = cur_list
            break
    if exit_mtm_pts is None:
        # square off at scheduled exit
        exits_at_trigger = []
        mtm = 0.0
        for k, side, qty, e in entries:
            p = get_price(bars, k, side, x_min)
            if p is None: return None
            exits_at_trigger.append(p)
            if qty == -1: mtm += (e - p)
            else:         mtm += (p - e)
        exit_mtm_pts = mtm
    return {
        "entries": entries, "exits": exits_at_trigger,
        "credit_pts": credit, "mtm_pts": exit_mtm_pts,
        "pnl_rupees": exit_mtm_pts * lotsize * LOTS,
        "exit_minute": exit_m, "exit_reason": exit_reason,
    }

def hhmm(m):
    return f"{m//60:02d}:{m%60:02d}"

def build_trade_rows(chosen_map, cache):
    """chosen_map: dict[wd] -> trade-spec OR dict[(wd,sym)] -> trade-spec."""
    rows = []
    cum = 0.0
    # Iterate cache in chronological order
    for (d, sym) in sorted(cache.keys()):
        wd = d.strftime("%a")
        bars, dte, exp_s = cache[(d, sym)]
        # Find spec
        spec = None
        if (wd, sym) in chosen_map:
            spec = chosen_map[(wd, sym)]
        elif wd in chosen_map and chosen_map[wd].get("sym") == sym:
            spec = chosen_map[wd]
        if spec is None: continue
        atm = find_atm(bars, to_min(spec["entry"]), sym)
        if atm is None: continue
        legs = build_strategies(sym)[spec["strat"]](atm)
        det = simulate_detail(bars, legs, to_min(spec["entry"]),
                              to_min(spec["exit"]), spec["sl"], spec["tp"],
                              LOT_SIZE[sym])
        if det is None: continue
        cum += det["pnl_rupees"]
        # Format legs into columns
        leg_desc = []
        for (k, side, qty, e), x_px in zip(det["entries"], det["exits"]):
            sign = "SHORT" if qty == -1 else "LONG"
            leg_desc.append(f"{sign} {k}{side} en={e:.2f} ex={x_px:.2f}")
        rows.append({
            "date": d.isoformat(),
            "weekday": wd,
            "index": sym,
            "expiry": exp_s,
            "DTE": dte,
            "strategy": spec["strat"],
            "ATM_strike": atm,
            "entry_time": spec["entry"],
            "scheduled_exit": spec["exit"],
            "actual_exit": hhmm(det["exit_minute"]),
            "exit_reason": det["exit_reason"],
            "credit_points": round(det["credit_pts"], 2),
            "credit_rupees": round(det["credit_pts"] * LOT_SIZE[sym] * LOTS, 2),
            "SL_pct_of_credit": spec["sl"],
            "TP_pct_of_credit": spec.get("tp") or "",
            "mtm_points": round(det["mtm_pts"], 2),
            "pnl_rupees": round(det["pnl_rupees"], 2),
            "cumulative_pnl": round(cum, 2),
            "leg1": leg_desc[0] if len(leg_desc) > 0 else "",
            "leg2": leg_desc[1] if len(leg_desc) > 1 else "",
            "leg3": leg_desc[2] if len(leg_desc) > 2 else "",
            "leg4": leg_desc[3] if len(leg_desc) > 3 else "",
            "lots": LOTS,
            "lot_size": LOT_SIZE[sym],
        })
    return rows

def monthly_summary(rows):
    by_m = defaultdict(lambda: {"pnl":0.0, "n":0, "w":0, "tp":0, "sl":0, "time":0,
                                "worst": 0, "best": 0})
    for r in rows:
        m = r["date"][:7]
        v = by_m[m]
        v["pnl"] += r["pnl_rupees"]; v["n"] += 1
        if r["pnl_rupees"] > 0: v["w"] += 1
        v[r["exit_reason"].lower()] = v.get(r["exit_reason"].lower(),0) + 1
        if r["pnl_rupees"] < v["worst"]: v["worst"] = r["pnl_rupees"]
        if r["pnl_rupees"] > v["best"]:  v["best"]  = r["pnl_rupees"]
    out = []
    g_pnl = g_n = g_w = 0
    for m in sorted(by_m):
        v = by_m[m]
        g_pnl += v["pnl"]; g_n += v["n"]; g_w += v["w"]
        out.append({
            "month": m, "trades": v["n"], "wins": v["w"],
            "win_pct": round(100*v["w"]/v["n"], 1) if v["n"] else 0,
            "pnl": round(v["pnl"], 2),
            "avg_per_trade": round(v["pnl"]/v["n"], 2) if v["n"] else 0,
            "worst_day": round(v["worst"], 2),
            "best_day":  round(v["best"], 2),
            "TP_exits": v.get("tp",0), "SL_exits": v.get("sl",0),
            "Time_exits": v.get("time",0),
        })
    out.append({"month":"TOTAL","trades":g_n,"wins":g_w,
                "win_pct":round(100*g_w/g_n,1) if g_n else 0,
                "pnl":round(g_pnl,2),
                "avg_per_trade":round(g_pnl/g_n,2) if g_n else 0,
                "worst_day":"", "best_day":"",
                "TP_exits":"","SL_exits":"","Time_exits":""})
    return out

# ----- Top-5 per (wd, sym) re-run from defined-risk results -----
def load_top5():
    """Quick re-sweep just for top-5 export. Reuses sweep engine but only
    persists per-(wd,sym) top-5 by cum, with post-regime filter."""
    from strategy_definedrisk_sweep import (ENTRIES, EXITS, SL_PCTS, TP_PCTS,
                                            MIN_SAMPLE, COVERAGE)
    near = find_days()
    cache = {}
    for (d, sym), exps in near.items():
        ex, exp_s = exps[0]
        bars = load_chain(d, sym, exp_s)
        if bars: cache[(d, sym)] = (bars, (ex-d).days, exp_s)
    by_bucket = defaultdict(list)
    for (d, sym), (bars, dte, _) in cache.items():
        by_bucket[(d.strftime("%a"), sym)].append((d, bars, dte))
    rows = []
    for (wd, sym), days in by_bucket.items():
        if len(days) < MIN_SAMPLE: continue
        strat_map = build_strategies(sym)
        lot = LOT_SIZE[sym]
        atm_cache = {}
        for d, bars, _ in days:
            for e in ENTRIES:
                atm_cache[(d, e)] = find_atm(bars, to_min(e), sym)
        bucket_results = []
        for strat_name, leg_fn in strat_map.items():
            for e in ENTRIES:
                em = to_min(e)
                for x in EXITS:
                    xm = to_min(x)
                    if xm <= em: continue
                    for sl in SL_PCTS:
                        for tp in TP_PCTS:
                            pnls = []
                            for d, bars, _ in days:
                                atm = atm_cache.get((d, e))
                                if atm is None: continue
                                legs = leg_fn(atm)
                                r = simulate(bars, legs, em, xm, sl, tp, lot)
                                if r is None: continue
                                pnls.append((d, r[0], r[2]))
                            n = len(pnls)
                            if n < len(days) * COVERAGE: continue
                            cum = sum(p for _,p,_ in pnls)
                            w = sum(1 for _,p,_ in pnls if p > 0)
                            post_n = sum(1 for d,_,_ in pnls if d >= REGIME_SPLIT)
                            post   = sum(p for d,p,_ in pnls if d >= REGIME_SPLIT)
                            if not (post_n >= 4 and post > 0): continue
                            bucket_results.append({
                                "weekday": wd, "index": sym, "strategy": strat_name,
                                "entry": e, "exit": x, "SL_pct": sl,
                                "TP_pct": tp or "",
                                "n": n, "wins": w,
                                "win_pct": round(100*w/n, 1),
                                "cum_pnl": round(cum, 2),
                                "post_pnl": round(post, 2),
                                "worst_day": round(min(p for _,p,_ in pnls), 2),
                                "best_day":  round(max(p for _,p,_ in pnls), 2),
                            })
        bucket_results.sort(key=lambda r: -r["cum_pnl"])
        rows.extend(bucket_results[:5])
    return rows, cache

# ----- main -----
def main():
    print("Loading schedule from /tmp/strategy_definedrisk.json ...")
    with open("/tmp/strategy_definedrisk.json") as f:
        sched = json.load(f)

    # Re-sweep top-5 (also gives us a fresh cache)
    print("Re-sweeping for Top-5 and trade logs (this takes a while)...")
    top5_rows, cache = load_top5()
    print(f"  cache={len(cache)} days, top5_rows={len(top5_rows)}")

    # Phase-3a rows
    chosen_3a = {}
    for wd, spec in sched["phase3a"].items():
        chosen_3a[wd] = {"sym": spec["sym"], "strat": spec["strat"],
                          "entry": spec["entry"], "exit": spec["exit"],
                          "sl": int(spec["sl"]),
                          "tp": (int(spec["tp"]) if spec.get("tp") not in (None,"","None") else None)}
    rows_3a = build_trade_rows(chosen_3a, cache)
    monthly_3a = monthly_summary(rows_3a)

    # Phase-3b rows
    chosen_3b = {}
    for key, spec in sched["phase3b"].items():
        wd, sym = key.split("_")
        chosen_3b[(wd, sym)] = {"sym": sym, "strat": spec["strat"],
                                 "entry": spec["entry"], "exit": spec["exit"],
                                 "sl": int(spec["sl"]),
                                 "tp": (int(spec["tp"]) if spec.get("tp") not in (None,"","None") else None)}
    rows_3b = build_trade_rows(chosen_3b, cache)
    monthly_3b = monthly_summary(rows_3b)

    # Schedule definition sheet
    sched_def = []
    for wd in ["Mon","Tue","Wed","Thu","Fri"]:
        s = chosen_3a.get(wd)
        if s:
            sched_def.append({"phase":"3a", "weekday":wd, "index":s["sym"],
                              "strategy":s["strat"], "entry_time":s["entry"],
                              "exit_time":s["exit"], "SL_pct":s["sl"],
                              "TP_pct":s["tp"] or "none",
                              "lots":LOTS, "lot_size":LOT_SIZE[s["sym"]]})
    for (wd, sym), s in chosen_3b.items():
        sched_def.append({"phase":"3b", "weekday":wd, "index":sym,
                          "strategy":s["strat"], "entry_time":s["entry"],
                          "exit_time":s["exit"], "SL_pct":s["sl"],
                          "TP_pct":s["tp"] or "none",
                          "lots":LOTS, "lot_size":LOT_SIZE[sym]})

    # Write Excel
    print(f"\nWriting {OUT} ...")
    try:
        from openpyxl import Workbook
    except ImportError:
        import subprocess
        subprocess.run(["pip","install","openpyxl","--quiet"], check=True)
        from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active; ws.title = "Schedule_Definition"
    def write_sheet(ws, rows):
        if not rows:
            ws.append(["no data"]); return
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h,"") for h in headers])
        # autosize
        for i, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(r.get(h,""))) for r in rows])
            ws.column_dimensions[chr(64+i) if i<=26 else 'A'+chr(64+i-26)].width = min(max_len+2, 40)

    write_sheet(ws, sched_def)
    write_sheet(wb.create_sheet("Phase3a_Trades"), rows_3a)
    write_sheet(wb.create_sheet("Phase3a_Monthly"), monthly_3a)
    write_sheet(wb.create_sheet("Phase3b_Trades"), rows_3b)
    write_sheet(wb.create_sheet("Phase3b_Monthly"), monthly_3b)
    write_sheet(wb.create_sheet("Weekday_Top5"), top5_rows)
    wb.save(OUT)

    print(f"\nDone. Sheets:")
    print(f"  Schedule_Definition  ({len(sched_def)} rows)")
    print(f"  Phase3a_Trades       ({len(rows_3a)} rows)  cum=₹{rows_3a[-1]['cumulative_pnl'] if rows_3a else 0:,.0f}")
    print(f"  Phase3a_Monthly      ({len(monthly_3a)} rows)")
    print(f"  Phase3b_Trades       ({len(rows_3b)} rows)  cum=₹{rows_3b[-1]['cumulative_pnl'] if rows_3b else 0:,.0f}")
    print(f"  Phase3b_Monthly      ({len(monthly_3b)} rows)")
    print(f"  Weekday_Top5         ({len(top5_rows)} rows)")
    print(f"\nFile: {OUT}")

if __name__ == "__main__":
    main()
